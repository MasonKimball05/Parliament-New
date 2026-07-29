"""
Public member directory view.
Shows basic member information visible to all authenticated members.
"""
import csv
import io
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from src.models import ParliamentUser


@login_required
def member_directory(request):
    """Display a public directory of all active members."""
    show_alumni = request.GET.get('show_alumni') == '1'

    # v3.17.3: directory.html renders each member's role badges, so every one of
    # these querysets needs `roles` prefetched — dev mode reported 12 + 11
    # repeats of the two role queries on a single directory load, attributed to
    # `directory.html:128 {% if member.roles.exists %}` and `:130 {% for role in
    # member.roles.all %}`.
    #
    # The template also had to change: `.exists()` on a prefetched relation
    # **bypasses the prefetch cache** and issues its own query, which is why
    # prefetching alone would have fixed only half of it. That is item five on
    # CLAUDE.md's own performance-sweep checklist, and this is what it looks like
    # in the wild.
    members = ParliamentUser.objects.filter(
        member_status='Active'
    ).exclude(
        member_type='Advisor'
    ).prefetch_related('roles').order_by('name')

    # Get advisors separately (materialized to list so .sort() works in sort branches below)
    advisors = list(ParliamentUser.objects.filter(
        member_status='Active',
        member_type='Advisor'
    ).prefetch_related('roles').order_by('name'))

    alumni = []
    if show_alumni:
        alumni = list(ParliamentUser.objects.filter(
            member_status='Alumni'
        ).prefetch_related('roles').order_by('name'))

    # Group members by type for display
    officers = [m for m in members if m.member_type == 'Officer']
    chairs = [m for m in members if m.member_type == 'Chair']
    regular_members = [m for m in members if m.member_type == 'Member']
    pledges = [m for m in members if m.member_type == 'Pledge']

    sort = request.GET.get('sort', 'name')
    if sort == 'roll':
        for group in [officers, chairs, regular_members, advisors]:
            group.sort(key=lambda m: (m.role_number is None, m.role_number or 0))
    elif sort == 'name_desc':
        for group in [officers, chairs, regular_members, pledges, advisors, alumni]:
            group.sort(key=lambda m: m.name, reverse=True)
    # default 'name' is already sorted ascending from the queryset

    def _can_set_house(user):
        if user.member_type == 'Officer' or user.is_admin:
            return True
        if user.member_type == 'Chair':
            return user.roles.filter(name__icontains='historian').exists()
        return False

    context = {
        'officers': officers,
        'chairs': chairs,
        'members': regular_members,
        'pledges': pledges,
        'advisors': advisors,
        'alumni': alumni,
        'show_alumni': show_alumni,
        # members queryset is already evaluated into the four lists above;
        # calling .count() would fire a fresh SQL COUNT — use len() instead.
        'total_count': len(officers) + len(chairs) + len(regular_members) + len(pledges) + len(advisors),
        'sort': sort,
        'filter_type': request.GET.get('filter', 'all'),
        'can_set_house': _can_set_house(request.user),
        'house_choices': ParliamentUser.HOUSE_CHOICES,
    }

    return render(request, 'directory.html', context)


@login_required
def export_directory(request):
    """Export the member directory in various formats (CSV, TXT, XLSX)."""
    format_type = request.GET.get('format', 'csv').lower()

    if request.GET.get('_export_submitted'):
        # Form was explicitly submitted — treat checkbox absence as unchecked
        include_alumni = 'include_alumni' in request.GET
        include_roll_number = 'include_roll_number' in request.GET
        include_user_id = 'include_user_id' in request.GET
        include_email = 'include_email' in request.GET
        include_phone = 'include_phone' in request.GET
        include_member_type = 'include_member_type' in request.GET
        include_member_status = 'include_member_status' in request.GET
        include_roles = 'include_roles' in request.GET
    else:
        # Direct link — sensible defaults
        include_alumni = False
        include_roll_number = True
        include_user_id = False
        include_email = True
        include_phone = True
        include_member_type = True
        include_member_status = False
        include_roles = True

    # Build column spec: list of (header_label, row_key)
    columns = [('Name', 'name')]
    if include_roll_number:
        columns.append(('Roll Number', 'roll_number'))
    if include_user_id:
        columns.append(('User ID', 'user_id'))
    if include_email:
        columns.append(('Email', 'email'))
    if include_phone:
        columns.append(('Phone', 'phone'))
    if include_member_type:
        columns.append(('Member Type', 'member_type'))
    if include_member_status:
        columns.append(('Status', 'member_status'))
    if include_roles:
        columns.append(('Role(s)', 'roles'))

    statuses = ['Active']
    if include_alumni:
        statuses.append('Alumni')

    all_members = ParliamentUser.objects.filter(
        member_status__in=statuses
    ).order_by('member_type', 'name').prefetch_related('roles')

    rows = []
    for member in all_members:
        # .exists() bypasses the prefetch_related cache and fires an extra DB query
        # per member. Use list(.all()) instead — it hits the prefetch cache.
        _member_roles = list(member.roles.all()) if include_roles else []
        roles_str = ', '.join(r.name for r in _member_roles)
        rows.append({
            'name': member.name,
            'roll_number': str(member.role_number) if member.role_number else '',
            'user_id': str(member.user_id),
            'email': member.email or '',
            'phone': member.phone_number or '',
            'member_type': member.member_type,
            'member_status': member.member_status,
            'roles': roles_str,
        })

    if format_type == 'txt':
        return _export_txt(rows, columns)
    elif format_type in ['xlsx', 'excel']:
        return _export_xlsx(rows, columns)
    else:
        return _export_csv(rows, columns)


def _export_csv(rows, columns):
    """Export directory as CSV file."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="member_directory.csv"'

    writer = csv.writer(response)
    writer.writerow([col[0] for col in columns])

    for row in rows:
        writer.writerow([row.get(col[1], '') for col in columns])

    return response


def _export_txt(rows, columns):
    """Export directory as plain text file."""
    response = HttpResponse(content_type='text/plain')
    response['Content-Disposition'] = 'attachment; filename="member_directory.txt"'

    lines = ['MEMBER DIRECTORY', '=' * 60, '']

    for row in rows:
        for label, key in columns:
            value = row.get(key, '')
            if value:
                lines.append(f"{label}: {value}")
        lines.append('-' * 40)

    response.write('\n'.join(lines))
    return response


def _export_xlsx(rows, columns):
    """Export directory as Excel file."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return _export_csv(rows, columns)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Member Directory"

    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')

    for col_idx, (label, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row_num, row in enumerate(rows, 2):
        for col_idx, (_, key) in enumerate(columns, 1):
            ws.cell(row=row_num, column=col_idx, value=row.get(key, ''))

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="member_directory.xlsx"'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.read())

    return response
